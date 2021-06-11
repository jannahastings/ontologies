import os
from pathlib import Path
import re
import openpyxl
from openpyxl.styles import Font
import csv
import codecs
import sys
import argparse
import subprocess
import io

def add_extra_values(header, row, aggregate):
    print("aggregate is: ", aggregate)
    aggregate_list = aggregate.split(";")
    extra_rows = []

    for agg in aggregate_list:
    # for r in range(4): #todo: use aggregate values, insert correct cell values
        extra_values = {}
        for key, cell in zip(header, row):
            extra_values[key] = cell.value
        if any(extra_values.values()):
            extra_rows.append(extra_values)
            # print("extra_values: ", extra_values.values())
    return extra_rows


## PROGRAM EXECUTION --- required argument: input file name
if __name__ == '__main__':

    parser=argparse.ArgumentParser()
    parser.add_argument('--inputExcel', '-i',help='Name of the input Excel spreadsheet file')
    
    args=parser.parse_args()

    inputFileName = args.inputExcel
    
    if inputFileName is None :
        parser.print_help()
        sys.exit('Not enough arguments. Expected at least -i "Excel file name" ')

    pathpath = str(Path(inputFileName).parents[0])
    basename = str(Path(inputFileName).stem)
    suffix = str(Path(inputFileName).suffix)
    
    wb = openpyxl.load_workbook(inputFileName) #call with full path and filename? Much better
    sheet = wb.active
    data = sheet.rows
    rows = []
    aggregate_list = ["Mean", "Minimum", "Maximum", "Median"]
    header = [i.value for i in next(data)]
    # print("got header: ", header)
    for row in sheet[2:sheet.max_row]:
        values = {}
        extra_rows = []
        for key, cell in zip(header, row):
            values[key] = cell.value
            if key == "Aggregate" and cell.value != None:
                extra_rows = add_extra_values(header, row, cell.value)
        if any(values.values()):
            rows.append(values)
            # print("values: ", values.values())
        for extra_row in extra_rows:
            print("got extra row")
            rows.append(extra_row)
        
            
        
    
    # print("reached rows ", rows)
    for r in range(len(rows)):
        row = [v for v in rows[r].values()]
        if "Aggregate" in header:
            cell = row[header.index("Aggregate")]
            # print("cell is: ", cell)
            # rows.insert(3, row)
            # rows.insert(r, row) #test insert
            #todo: split by ";" and create new rows
    

    #new sheet to save:

    save_wb = openpyxl.Workbook()
    save_sheet = save_wb.active

    for c in range(len(header)):
        save_sheet.cell(row=1, column=c+1).value=header[c]
        save_sheet.cell(row=1, column=c+1).font = Font(size=12,bold=True)
    for r in range(len(rows)):
        row = [v for v in rows[r].values()]
        for c in range(len(header)):
            save_sheet.cell(row=r+2, column=c+1).value=row[c]
            
    #save:   
    save_wb.save(pathpath + "/" + basename + "_Expanded.xlsx")
    

